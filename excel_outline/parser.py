from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def parse_workbook(path: str | Path) -> dict[str, Any]:
    workbook_path = Path(path).expanduser().resolve()

    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Only .xlsx files are supported: {workbook_path}")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False)

    sheets_out: list[dict[str, Any]] = []
    known_maps: dict[str, dict[str, dict[str, Any]]] = {}
    parsed_sheet_names: set[str] = set()

    # pass 1 — detect map sheets
    for ws in wb.worksheets:
        map_element = detect_map_sheet(ws)
        if map_element is None:
            continue

        sheets_out.append(
            {
                "name": ws.title,
                "elements": [map_element],
            }
        )
        parsed_sheet_names.add(ws.title)

        map_lookup: dict[str, dict[str, Any]] = {}

        for key, entry in map_element["entries"].items():
            value_anchor = entry["value_anchor"]
            value_cell = value_anchor.split("!", 1)[1]

            map_lookup[key] = {
                "anchor": entry["anchor"],
                "value_anchor": value_anchor,
                "value_cell": value_cell,
                "value": entry["value"],
            }

        known_maps[ws.title] = map_lookup

    # pass 2 — detect table sheets
    for ws in wb.worksheets:
        if ws.title in parsed_sheet_names:
            continue

        table_element = detect_table_sheet(ws, known_maps)

        if table_element is not None:
            sheets_out.append(
                {
                    "name": ws.title,
                    "elements": [table_element],
                }
            )
        else:
            sheets_out.append(
                {
                    "name": ws.title,
                    "elements": [],
                }
            )

    return {"sheets": sheets_out}


def detect_map_sheet(ws: Worksheet) -> dict[str, Any] | None:
    if ws.max_row < 2 or ws.max_column < 2:
        return None

    if ws["A1"].value is None or ws["B1"].value is None:
        return None

    entries: dict[str, dict[str, Any]] = {}

    for row in range(2, ws.max_row + 1):
        key_cell = ws[f"A{row}"]
        value_cell = ws[f"B{row}"]

        key = key_cell.value
        value = value_cell.value

        if key is None and value is None:
            continue

        if not isinstance(key, str) or key.strip() == "":
            return None

        if isinstance(value, str) and value.startswith("="):
            return None

        entries[key] = {
            "anchor": make_anchor(ws.title, key_cell.coordinate),
            "value_anchor": make_anchor(ws.title, value_cell.coordinate),
            "value": value,
        }

    if not entries:
        return None

    return {
        "kind": "map",
        "anchor": make_anchor(ws.title, "A1"),
        "entries": entries,
    }


def detect_table_sheet(
    ws: Worksheet,
    known_maps: dict[str, dict[str, dict[str, Any]]],
) -> dict[str, Any] | None:

    if ws.max_row < 2 or ws.max_column < 3:
        return None

    if ws["A1"].value is None:
        return None

    header_values = [
        ws.cell(row=1, column=col).value
        for col in range(2, ws.max_column + 1)
    ]

    # detect header range like 1..10
    columns_info = detect_integer_range(
        ws.title,
        2,
        ws.max_column,
        header_values,
    )

    if columns_info is None:
        return None

    series: dict[str, dict[str, Any]] = {}

    for row in range(2, ws.max_row + 1):

        label = ws.cell(row=row, column=1).value

        if not isinstance(label, str) or label.strip() == "":
            return None

        formula_cells = [
            ws.cell(row=row, column=col)
            for col in range(2, ws.max_column + 1)
        ]

        if not all(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for cell in formula_cells
        ):
            return None

        rule_info = infer_series_rule(
            ws,
            row,
            2,
            ws.max_column,
            known_maps,
        )

        if rule_info is None:
            return None

        start_coord = ws.cell(row=row, column=2).coordinate
        end_coord = ws.cell(row=row, column=ws.max_column).coordinate

        series[label] = {
            "anchor": make_anchor(ws.title, f"A{row}"),
            "range": f"{ws.title}!{start_coord}:{end_coord}",
            "rule": rule_info["rule"],
            "depends_on": rule_info["depends_on"],
        }

    return {
        "kind": "table",
        "anchor": make_anchor(ws.title, "A1"),
        "columns": columns_info,
        "series": series,
    }


def detect_integer_range(
    sheet_name: str,
    start_col: int,
    end_col: int,
    values: list[Any],
) -> dict[str, Any] | None:

    if not values:
        return None

    if not all(isinstance(v, int) for v in values):
        return None

    if len(values) == 1:
        step = 1
    else:
        step = values[1] - values[0]

        if step == 0:
            return None

        for i in range(2, len(values)):
            if values[i] - values[i - 1] != step:
                return None

    start_coord = f"{get_column_letter(start_col)}1"
    end_coord = f"{get_column_letter(end_col)}1"

    return {
        "anchor": make_anchor(sheet_name, start_coord),
        "range": f"{sheet_name}!{start_coord}:{end_coord}",
        "kind": "integer_range",
        "start": values[0],
        "end": values[-1],
        "step": step,
    }


def infer_series_rule(
    ws: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    known_maps: dict[str, dict[str, dict[str, Any]]],
) -> dict[str, Any] | None:

    formulas: list[str] = []

    for col in range(start_col, end_col + 1):
        value = ws.cell(row=row, column=col).value

        if not isinstance(value, str) or not value.startswith("="):
            return None

        formulas.append(value)

    # try known rule patterns
    rule = try_infer_column_times_variable(
        ws,
        row,
        start_col,
        end_col,
        formulas,
        known_maps,
    )

    if rule is not None:
        return rule

    rule = try_infer_row_sum(
        ws,
        row,
        start_col,
        end_col,
        formulas,
    )

    return rule


def try_infer_column_times_variable(
    ws: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    formulas: list[str],
    known_maps: dict[str, dict[str, dict[str, Any]]],
) -> dict[str, Any] | None:

    fixed_variable: str | None = None
    fixed_sheet: str | None = None

    for offset, col in enumerate(range(start_col, end_col + 1)):

        formula = formulas[offset]
        body = formula[1:]
        parts = body.split("*")

        if len(parts) != 2:
            return None

        left, right = parts[0].strip(), parts[1].strip()

        expected_a = f"${get_column_letter(col)}$1"
        expected_b = f"{get_column_letter(col)}$1"

        if left in (expected_a, expected_b):
            other = right
        elif right in (expected_a, expected_b):
            other = left
        else:
            return None

        normalized = normalize_excel_ref(other)

        lookup = lookup_variable_by_value_ref(
            normalized,
            known_maps,
        )

        if lookup is None:
            return None

        sheet_name, variable_name = lookup

        if fixed_variable is None:
            fixed_variable = variable_name
            fixed_sheet = sheet_name
        elif (
            variable_name != fixed_variable
            or sheet_name != fixed_sheet
        ):
            return None

    return {
        "rule": f"column * {fixed_variable}",
        "depends_on": [f"{fixed_sheet}.{fixed_variable}", "column"],
    }


def try_infer_row_sum(
    ws: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    formulas: list[str],
) -> dict[str, Any] | None:

    reference_rows: tuple[int, int] | None = None

    for offset, col in enumerate(range(start_col, end_col + 1)):

        formula = formulas[offset]
        body = formula[1:]
        parts = body.split("+")

        if len(parts) != 2:
            return None

        left, right = parts[0].strip(), parts[1].strip()

        expected_col = get_column_letter(col)

        left_col, left_row = split_a1_ref(
            normalize_excel_ref(left)
        )

        right_col, right_row = split_a1_ref(
            normalize_excel_ref(right)
        )

        if left_col != expected_col or right_col != expected_col:
            return None

        pair = (left_row, right_row)

        if reference_rows is None:
            reference_rows = pair
        elif reference_rows != pair:
            return None

    row1, row2 = reference_rows

    label1 = ws.cell(row=row1, column=1).value
    label2 = ws.cell(row=row2, column=1).value

    if not isinstance(label1, str) or not isinstance(label2, str):
        return None

    return {
        "rule": f"{label1} + {label2}",
        "depends_on": [label1, label2],
    }


def lookup_variable_by_value_ref(
    normalized_ref: str,
    known_maps: dict[str, dict[str, dict[str, Any]]],
) -> tuple[str, str] | None:

    if "!" not in normalized_ref:
        return None

    sheet_name, cell_ref = normalized_ref.split("!", 1)

    if sheet_name not in known_maps:
        return None

    for name, info in known_maps[sheet_name].items():
        if info["value_cell"] == cell_ref:
            return sheet_name, name

    return None


def normalize_excel_ref(ref: str) -> str:
    return ref.replace("$", "")


def split_a1_ref(ref: str) -> tuple[str, int]:

    letters: list[str] = []
    digits: list[str] = []

    for c in ref:
        if c.isalpha():
            letters.append(c)
        elif c.isdigit():
            digits.append(c)

    if not letters or not digits:
        raise ValueError(ref)

    return "".join(letters), int("".join(digits))


def make_anchor(sheet_name: str, coordinate: str) -> str:
    return f"{sheet_name}!{coordinate}"
