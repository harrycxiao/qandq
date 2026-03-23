# Make sure to set:
# env:OPENAI_API_KEY=your_api_key_here
# before running this script.

from pydantic import BaseModel
from pydantic_ai import Agent
from openpyxl import load_workbook

from typing import Literal, Optional
import sys

# ---------- Excel helpers ----------

def list_sheets(path: str):
    wb = load_workbook(path, read_only=True)
    return wb.sheetnames


def preview(path: str, sheet: str, n: int):
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]

    rows = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n:
            break
        rows.append(list(row))

    return rows


# ---------- Tool schema ----------

class Tool(BaseModel):
    tool: Literal["list_sheets", "preview"]
    sheet: Optional[str] = None
    n: Optional[int] = None


# ---------- Agent ----------

agent = Agent(
    model="openai:gpt-4.1",
    output_type=Tool,
    system_prompt="""
You are an agent with two tools to choose from,
and you must choose exactly one.

Tools:

list_sheets:
    returns the sheet names in the workbook

preview:
    returns the first n rows of a given sheet

Decide which tool to call based on the user request.
""",
)


# ---------- CLI ----------

def cli():
    if len(sys.argv) < 3:
        print("Usage:")
        print('python cli.py <workbook_path> "<question>"')
        sys.exit(1)

    file = sys.argv[1]
    question = sys.argv[2]

    decision = agent.run_sync(question).output

    if decision.tool == "list_sheets":

        result = list_sheets(file)

    elif decision.tool == "preview":

        if decision.sheet is None:
            raise ValueError("preview requires sheet name")

        if decision.n is None:
            raise ValueError("preview requires n")

        result = preview(file, decision.sheet, decision.n)

    else:
        raise ValueError(f"Unknown tool: {decision.tool}")

    print(result)


# ---------- entry point ----------

if __name__ == "__main__":
    cli()
